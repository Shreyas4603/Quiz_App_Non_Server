import tkinter
from tkinter import *
import customtkinter as ctk
from customtkinter import CTkToplevel
import time
import customtkinter as ct
from tkinter import filedialog, END, ACTIVE
import shutil

from pyglet import *
import loadFont
import openpyxl
from openpyxl.utils import get_column_letter

# configure
bg = "#1a1a1c"
accent = "#07d969"
text = "#fff6df"
secondary = "#c3083f"

root = ctk.CTk()

width = 700
height = 500
root.geometry(str(width) + "x" + str(height))
root.maxsize(width, height)
root.minsize(width, height)
root.configure(fg_color=bg)
root.title("Quiz Master 3000")

f_path = os.getcwd() + "\Fonts\\"

fonts = [f_path + "Momcake", f_path + "Adobe Gothic Std B", f_path + "Bebas Neue"]
loadFont.loadfont(fonts[0])
loadFont.loadfont(fonts[1])
loadFont.loadfont(fonts[2])

# fonts=font.load("Ubuntu Bold")

# configure

# global variable start


filename_arr = [""]
profile_name = ct.StringVar()
question = []
A = []
B = []
C = []
D = []
ref = [A, B, C, D]
correct = []
ansSelected = []
questionCount = 0


# global variable end


# ---------------------------------------------------------------------------------------------------------------


# Main Fuction Start


def getExcelData(excelBook):  # takes workbook object

    questionCount = 0
    ws = excelBook.active
    r = ws.rows
    checkHead = ['questions', 'question', 'a', 'b', 'c', 'd', 'answers', 'answer', '']
    num_rows = len(ws['A'])
    flag = 0
    for i in ws['A']:
        if (str(i.value).strip().lower() in checkHead[0:2]):
            flag = 1
            continue

        elif (str(i.value).strip().lower() != "none" and flag == 1):
            flag = 1
            question.append(str(i.value).strip())

    for row in range(num_rows):
        idx = 0
        flag = 0
        for col in range(2, 6):

            let = get_column_letter(col)

            if (str(ws[let + str(row + 1)].value).strip().lower() in checkHead[2:6]):
                flag = 1

                continue

            elif (str(ws[let + str(row + 1)].value).strip().lower() != "none"):
                flag = 1

                ref[idx].append(str(ws[let + str(row + 1)].value).strip())
                idx += 1
    print(ref)

    flag = 0
    for i in ws['F']:
        if (str(i.value).strip().lower() in checkHead[6:8]):
            flag = 1
            continue

        elif (str(i.value).strip().lower() != "none" and flag == 1):
            flag = 1
            correct.append(str(i.value).strip())


#
#
# 74 chars length
# getExcelData(openpyxl.load_workbook("E:\Quiz App\Questions\Question.xlsx"))#for debug delete later\
c = 1


def results():
    print(ansSelected)

    correctAns = 0

    for i, j in zip(ansSelected, correct):
        print('me : ', i, 'right : ', j)
        if (i == j):
            correctAns += 1
    print("Number of answers correct : ", correctAns)

    result = str(correctAns) + '/' + str(len(correct))

    resultWin = CTkToplevel(root)
    resultWin.geometry(str(450) + "x" + str(400))
    resultWin.maxsize(450, 400)
    resultWin.minsize(450, 400)
    resultWin.configure(fg_color=bg)

    resultHead = ct.CTkLabel(master=resultWin,
                             text="Results", font=("Momcake", 60, "bold"), fg_color=secondary,
                             bg_color="transparent",
                             text_color=text,
                             corner_radius=8)
    resultHead.pack(pady=(20, 20))

    resultDis = ct.CTkLabel(master=resultWin,
                            text=result, font=("Bebas Neue", 100, "bold"),
                            bg_color="transparent",
                            text_color=text,
                            corner_radius=8)
    resultDis.pack()

    resultText = ct.CTkLabel(master=resultWin,
                             text="Correct", font=("Bebas Neue", 50, "bold"),
                             bg_color="transparent",
                             text_color=accent,
                             corner_radius=8)
    resultText.pack()


def nextLoad():
    global questionCount

    questionCount += 1
    if questionCount == len(question):
        nextQuestion.configure(text='Results')
        ansSelected.append(str(option_selected.get()))
        questionCount = 0
        results()
        quizStart.destroy()

        return

    ansSelected.append(str(option_selected.get()))

    print(' in next count ', questionCount)
    sentence = question[questionCount]
    sentence, h = processData(sentence)

    a, b, c, d = A[questionCount], B[questionCount], C[questionCount], D[questionCount]

    questionLable.configure(text=sentence, height=h)
    optionA.deselect()
    optionA.configure(text=a, variable=option_selected)
    optionB.configure(text=b, variable=option_selected)
    optionC.configure(text=c, variable=option_selected)
    optionD.configure(text=d, variable=option_selected)


def processData(que):  # calculate the len of the question and determin the rquired label size
    s = ""
    l = 1
    for idx, i in enumerate(que, start=1):

        if idx % 74 != 0:
            s = s + i
        else:
            print('in 28', idx)
            s = s + i + '\n'
            l += 1
    print(s)
    return (s, l * 30)


def startMain():
    name = profile_entry.get()
    startWin.destroy()
    if (name not in profList):
        enter_p_name.configure(text="Profile doesn't exist", text_color=secondary)
        global excelPathSend
    try:
        excelPathSend = os.getcwd() + '\\Profiles\\' + name + '.xlx'
        wb = openpyxl.load_workbook(excelPathSend)

    except:
        excelPathSend = os.getcwd() + '\\Profiles\\' + name + '.xlsx'
        wb = openpyxl.load_workbook(excelPathSend)
    getExcelData(wb)
    print("final question = ", question, len(question))

    # window creation
    global quizStart
    quizStart = CTkToplevel(root)
    quizStart.maxsize(width, 500)
    quizStart.geometry(str(width) + "x" + str(500))
    quizStart.minsize(width, 500)
    quizStart.title('Quiz Master 3000')
    quizStart.configure(fg_color=bg)

    sentence = question[questionCount]
    sentence, h = processData(sentence)

    global questionLable
    questionLable = ct.CTkLabel(quizStart, text=sentence, font=("Gill Sans MT", 20), justify='center',
                                fg_color=secondary, width=width, height=h)
    questionLable.pack(pady=(0, 30))

    global option_selected, optionA, optionB, optionC, optionD
    option_selected = ct.StringVar()
    a, b, c, d = A[questionCount], B[questionCount], C[questionCount], D[questionCount]
    optionA = ct.CTkRadioButton(quizStart, text=a, font=("Gill Sans MT", 20), variable=option_selected, value='A',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionB = ct.CTkRadioButton(quizStart, text=b, font=("Gill Sans MT", 20), variable=option_selected, value='B',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionC = ct.CTkRadioButton(quizStart, text=c, font=("Gill Sans MT", 20), variable=option_selected, value='C',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionD = ct.CTkRadioButton(quizStart, text=d, font=("Gill Sans MT", 20), variable=option_selected, value='D',
                                border_color=secondary, fg_color=accent, hover_color=accent)

    # optionA.pack(pady=(50,30))
    # optionA.grid(row=1,column=0)

    optionA.place(x=200, y=150)
    optionB.place(x=200, y=200)
    optionC.place(x=200, y=250)
    optionD.place(x=200, y=300)

    global nextQuestion
    nextQuestion = ctk.CTkButton(quizStart, text="Next", font=("Adobe Gothic Std B", 30), width=100, fg_color=bg,
                                 text_color=text, hover_color=secondary, border_color=secondary, border_width=2,
                                 corner_radius=50, command=nextLoad)
    nextQuestion.place(x=294, y=381)

    quizStart.mainloop()


def quickMain():
    print('in quixk main')
    print('fila name', filename_arr)
    questionCount = 0

    create_window.destroy()

    # name = profile_entry.get()
    #
    #
    # if (name not in profList):
    #     enter_p_name.configure(text="Profile doesn't exist", text_color=secondary)
    #     global excelPathSend
    # try:
    #     excelPathSend = os.getcwd() + '\\Profiles\\' + name + '.xlx'
    #     wb = openpyxl.load_workbook(excelPathSend)
    #
    # except:
    #     excelPathSend = os.getcwd() + '\\Profiles\\' + name + '.xlsx'
    wb = openpyxl.load_workbook(filename_arr[0])
    getExcelData(wb)
    print("final question = ", question, len(question))

    # window creation
    global quizStart
    quizStart = CTkToplevel(root)
    quizStart.maxsize(width, 500)
    quizStart.geometry(str(width) + "x" + str(500))
    quizStart.minsize(width, 500)
    quizStart.title('Quiz Master 3000')
    quizStart.configure(fg_color=bg)

    sentence = question[questionCount]
    sentence, h = processData(sentence)

    global questionLable
    questionLable = ct.CTkLabel(quizStart, text=sentence, font=("Gill Sans MT", 20), justify='center',
                                fg_color=secondary, width=width, height=h)
    questionLable.pack(pady=(0, 30))

    global option_selected, optionA, optionB, optionC, optionD
    option_selected = ct.StringVar()
    a, b, c, d = A[questionCount], B[questionCount], C[questionCount], D[questionCount]
    optionA = ct.CTkRadioButton(quizStart, text=a, font=("Gill Sans MT", 20), variable=option_selected, value='A',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionB = ct.CTkRadioButton(quizStart, text=b, font=("Gill Sans MT", 20), variable=option_selected, value='B',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionC = ct.CTkRadioButton(quizStart, text=c, font=("Gill Sans MT", 20), variable=option_selected, value='C',
                                border_color=secondary, fg_color=accent, hover_color=accent)
    optionD = ct.CTkRadioButton(quizStart, text=d, font=("Gill Sans MT", 20), variable=option_selected, value='D',
                                border_color=secondary, fg_color=accent, hover_color=accent)

    # optionA.pack(pady=(50,30))
    # optionA.grid(row=1,column=0)

    optionA.place(x=200, y=150)
    optionB.place(x=200, y=200)
    optionC.place(x=200, y=250)
    optionD.place(x=200, y=300)

    global nextQuestion
    nextQuestion = ctk.CTkButton(quizStart, text="Next", font=("Adobe Gothic Std B", 30), width=100, fg_color=bg,
                                 text_color=text, hover_color=secondary, border_color=secondary, border_width=2,
                                 corner_radius=50, command=nextLoad)
    nextQuestion.place(x=294, y=381)

    quizStart.mainloop()


# Main Fuction End

# ---------------------------------------------------------------------------------------------------------------


# side util functon Start

def check_folder():  # to check if questions and profile folder exist

    global p_path
    path = os.getcwd()  # current directory we r working in
    s_path = path + "\Questions"
    p_path = path + "\Profiles"
    if (os.path.exists(s_path) == False):
        os.mkdir(s_path)
        print("Created Questions ")
        return
    elif (os.path.exists(p_path) == False):
        print("Created Profiles ")
        os.mkdir(p_path)
        return
    else:
        print("exist")


def getProfileList():  # list of existing profiles
    path = os.getcwd() + "\Profiles"
    global profList
    profList = os.listdir(path)
    profList = [str(i.split(".")[0]) for i in profList]
    return (profList)


def updateListBox(data):
    # insert the profile names to the list
    profile_list.delete(0, END)
    for item in data[:9]:
        profile_list.insert(END, item)


def fillout(e):  # to fill the entry box on clicking the list item
    profile_entry.delete(0, END)

    profile_entry.insert(0, profile_list.get(ANCHOR))


def check(e):
    # grab what is typed
    typed = profile_entry.get()
    profiles = getProfileList()  ##store the profile names in list form
    if typed == "":
        data = profiles
    else:
        data = []
        for item in profiles:
            if typed.lower() in item.lower():
                data.append(item)
    updateListBox(data)


# side util functon End

# ---------------------------------------------------------------------------------------------------------------


# window creation button functions start

def start_quiz_function():
    # clear the values first
    question.clear()
    A.clear()
    B.clear()
    C.clear()
    D.clear()
    correct.clear()
    ansSelected.clear()
    # clear the values first

    global startWin  # creates start quiz landing window
    startWin = CTkToplevel(root)
    width = 450
    height = 350
    startWin.geometry("450x500")
    startWin.resizable(0, 0)
    startWin.title("Start quiz")

    startWin.configure(fg_color=bg)
    start_quiz = ct.CTkLabel(master=startWin,
                             text="Start Quiz", font=("Momcake", 60, "bold"), fg_color=secondary,
                             bg_color="transparent",
                             text_color=text,
                             corner_radius=8)
    start_quiz.pack(pady=10)  # start quiz title

    global enter_p_name
    enter_p_name = ct.CTkLabel(startWin, text="Enter profile name", font=("Gill Sans MT", 20))
    enter_p_name.place(x=148, y=106)

    profile_name = ct.StringVar()
    global profile_entry
    profile_entry = ct.CTkEntry(master=startWin, textvariable=profile_name, font=("Consolas", 20), width=300,
                                border_width=3, border_color=secondary, justify="left")
    profile_entry.pack(pady=(60, 10))

    global profile_list
    profile_list = tkinter.Listbox(startWin, width=26, height=9, font=("Consolas", 15), background=bg, foreground=text,
                                   selectbackground=secondary,
                                   borderwidth=3, highlightcolor=secondary, relief="flat",
                                   highlightbackground=secondary)
    profile_list.place(x=77, y=183)

    updateListBox(getProfileList())
    profile_list.bind("<<ListboxSelect>>", fillout)
    profile_entry.bind("<KeyRelease>", check)

    done = ct.CTkButton(master=startWin, text="Start", font=("Adobe Gothic Std B", 35), corner_radius=30, width=270,
                        command=startMain, fg_color="#2c8c00", hover_color="#2eab49")
    done.place(x=90, y=416)

    print("start quiz")


def quick_quiz_function():
    root.iconify()
    question.clear()
    correct.clear()
    A.clear()
    B.clear()
    C.clear()
    D.clear()
    ansSelected.clear()

    def open_file():
        global filename
        filename = ""
        filename = filename + str(filedialog.askopenfilename())
        filename_arr[0] = filename  # stores directory address of the question file selected
        # print(filename)
        # file=filedialog.askdirectory(title="Select image files")
        name = filename.split("/")[-1]
        print(name, " = name")
        file_name.configure(text=name)
        print("filanme", filename)

    global create_window
    create_window = CTkToplevel(root)
    width = 450
    height = 600
    create_window.geometry("450x350")
    create_window.resizable(0, 0)
    create_window.title("Quiz quiz")

    create_window.configure(fg_color=bg)
    create_quiz = ct.CTkLabel(master=create_window,
                              text="Quick Quiz", font=("Momcake", 60, "bold"), fg_color=secondary,
                              bg_color="transparent",
                              text_color=text,
                              corner_radius=8)
    create_quiz.pack(pady=10)  # create quiz title

    # enter_p_name = ct.CTkLabel(create_window, text="Enter profile name", font=("Gill Sans MT", 20))
    # enter_p_name.place(x=148, y=126)

    # profile_name = ct.StringVar()
    # profile_entry = ct.CTkEntry(master=create_window, textvariable=profile_name, font=("Consolas", 20), width=300,
    #                             border_width=3, border_color=secondary, justify="left")
    # profile_entry.pack(pady=(80, 10))

    select_file = ct.CTkButton(master=create_window, text="Open file", font=("Gill Sans MT", 20), corner_radius=5,
                               command=open_file,
                               fg_color=secondary, width=100)
    select_file.pack(pady=(50, 10))

    global file_name
    file_name = ct.CTkLabel(master=create_window, text="", font=("Gill Sans MT", 20), corner_radius=5, width=192,
                            height=35, text_color=text)
    file_name.pack(pady=(5, 10))

    done = ct.CTkButton(master=create_window, text="Start", font=("Gill Sans MT", 25), corner_radius=30, width=104,
                        command=quickMain, fg_color="#2c8c00", hover_color="#2eab49")
    done.pack(pady=(20, 10))


def create_quiz_function():
    root.iconify()

    def check_exist_profile():
        check_folder()
        global c  # to make sure file is selected ,flag var
        c = True
        print("p path", p_path)
        print("ext = ", ext)
        c = ext.isalpha()

        pth = p_path + "\\" + profile_name.get() + "." + ext
        print("Path profile check = ", pth)
        check = os.path.exists(pth)
        print("check = ", check)
        if c == False:
            print("c = ", c)
            file_name.configure(text="Select file")
        elif check == True:
            print("in check true")
            file_name.configure(text="Profile exist")
            create_window.update()
            time.sleep(1)
            return True
        return False

    def create_profile(prof_name):
        # print("name = ",prof_name.get())
        global ext, p_e

        # check_folder()  # check of question n profile exist
        try:
            source = filename_arr[0]
            ext = source.split(".")[-1]
            p_e = check_exist_profile()
            print("Profile exist ", p_e)
            if (p_e == False and c == True):
                f_name = source.split("Questions/")[-1]
                print(f_name)

                destination = os.getcwd() + "\Profiles\\"
                shutil.copy2(source, destination)

                src = destination + f_name
                dst = destination + str(prof_name.get()) + "." + ext

                print("src = ", src)
                print("dst = ", dst)
                os.rename(src, dst)



        except Exception as e:
            file_name.configure(text=e)
        if c == True and p_e == False:
            file_name.configure(text="Profile complete")
            create_window.update()
            time.sleep(1)
            root.state("zoomed")
            create_window.destroy()

    def open_file():
        global filename
        filename = ""
        filename = filename + str(filedialog.askopenfilename())
        filename_arr[0] = filename  # stores directory address of the question file selected
        # print(filename)
        # file=filedialog.askdirectory(title="Select image files")
        name = filename.split("/")[-1]
        print(name, " = name")
        file_name.configure(text=name)
        print(filename)

    global create_window
    create_window = CTkToplevel(root)
    width = 450
    height = 600
    create_window.geometry("450x350")
    create_window.resizable(0, 0)
    create_window.title("Create quiz")

    create_window.configure(fg_color=bg)
    create_quiz = ct.CTkLabel(master=create_window,
                              text="Create Quiz", font=("Momcake", 60, "bold"), fg_color=secondary,
                              bg_color="transparent",
                              text_color=text,
                              corner_radius=8)
    create_quiz.pack(pady=10)  # create quiz title

    enter_p_name = ct.CTkLabel(create_window, text="Enter profile name", font=("Gill Sans MT", 20))
    enter_p_name.place(x=148, y=126)

    profile_name = ct.StringVar()
    profile_entry = ct.CTkEntry(master=create_window, textvariable=profile_name, font=("Consolas", 20), width=300,
                                border_width=3, border_color=secondary, justify="left")
    profile_entry.pack(pady=(80, 10))

    select_file = ct.CTkButton(master=create_window, text="Open file", font=("Gill Sans MT", 20), corner_radius=5,
                               command=open_file,
                               fg_color=secondary, width=100)
    select_file.place(x=76, y=201)

    global file_name
    file_name = ct.CTkLabel(master=create_window, text="", font=("Gill Sans MT", 20), corner_radius=5, width=192,
                            height=35, text_color=text)
    file_name.place(x=180, y=201)

    done = ct.CTkButton(master=create_window, text="Create", font=("Gill Sans MT", 25), corner_radius=30, width=104,
                        command=lambda: create_profile(profile_name), fg_color="#2c8c00", hover_color="#2eab49")
    done.place(x=182, y=243)


# window creation button functions end

# ---------------------------------------------------------------------------------------------------------------


# Dev Tools Start

def dev_tools(atrib, root):
    global sl_x, sl_y, sl_w, sl_h
    # position slider
    slider_x = ctk.CTkSlider(root, from_=0, to=width, number_of_steps=width, command=slide_x)
    slider_x.place(x=100, y=580)

    slider_v = ctk.CTkSlider(root, from_=0, to=height, number_of_steps=height, command=slide_v,
                             orientation="vertical", )
    slider_v.place(x=100, y=400)

    sl_x = ctk.CTkLabel(root)
    sl_x.place(x=191, y=550)

    sl_y = ctk.CTkLabel(root)
    sl_y.place(x=40, y=490)

    # size slider
    slider_w = ctk.CTkSlider(root, from_=0, to=width, number_of_steps=width, command=size_w)
    slider_w.place(x=50, y=380)

    slider_h = ctk.CTkSlider(root, from_=0, to=height, number_of_steps=height, command=size_h, orientation="vertical", )
    slider_h.place(x=50, y=200)

    sl_w = ctk.CTkLabel(root)
    sl_w.place(x=191, y=350)

    sl_h = ctk.CTkLabel(root)
    sl_h.place(x=40, y=290)


def slide_x(value):
    atrib.place(x=int(value))
    sl_x.configure(text="x=" + str(int(value)))
    print("x=", value)


def slide_v(value):
    atrib.place(y=int(value))
    sl_y.configure(text="y=" + str(int(value)))
    print("y=", value)


def size_w(value):
    atrib.configure(width=int(value))
    sl_w.configure(text=int(value))


def size_h(value):
    atrib.configure(height=int(value))
    sl_h.configure(text=int(value))


# Dev Tools End

# ---------------------------------------------------------------------------------------------------------------


# main
check_folder()
welcome_text = ctk.CTkLabel(root, text="Welcome", font=("Bebas Neue", 100), text_color=secondary)
welcome_text.place(x=196, y=0)

tag_line = ctk.CTkLabel(root, text="Q u i z  M a s t e r  3 0 0 0", font=("MOMCAKE", 18), text_color=text)
tag_line.place(x=250, y=100)

create_quiz = ctk.CTkButton(root, text="Create Quiz", command=create_quiz_function, font=("Adobe Gothic Std B", 30),
                            fg_color=bg, text_color=text, hover_color=secondary, border_color=secondary, border_width=2,
                            corner_radius=50)
create_quiz.place(x=259, y=189)

start_quiz = ctk.CTkButton(root, text="Start   Quiz", font=("Adobe Gothic Std B", 30), width=190, fg_color=bg,
                           text_color=text, hover_color=secondary, border_color=secondary, border_width=2,
                           corner_radius=50, command=start_quiz_function)
start_quiz.place(x=259, y=248)

quick_quiz = ctk.CTkButton(root, text="Quick   Quiz", font=("Adobe Gothic Std B", 30), width=190, fg_color=bg,
                           text_color=text, hover_color=secondary, border_color=secondary, border_width=2,
                           corner_radius=50, command=quick_quiz_function)
quick_quiz.place(x=259, y=306)

# main
# dev tools - position and size
root.mainloop()
